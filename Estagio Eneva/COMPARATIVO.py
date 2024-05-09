from openpyxl import Workbook
from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl

def plot_grafico_comparativo(comparativo_por_estacao):
    # Esta função plota um gráfico comparativo do crescimento por estação ao longo dos anos
    anos = sorted(set([ano for estacao in comparativo_por_estacao.values() for ano in estacao.keys()]))
    estacoes = list(comparativo_por_estacao.keys())

    for estacao in estacoes:
        crescimento_por_ano = [comparativo_por_estacao[estacao].get(ano, 0) for ano in anos]
        plt.plot(anos, crescimento_por_ano, label=estacao)

    plt.xlabel('Ano')
    plt.ylabel('Crescimento')
    plt.title('Comparativo Anual por Estação')
    plt.legend()
    plt.grid(True)
    plt.show()

def main():
    # Função principal do programa
    comparativo_por_estacao = {
        'Verão': {2021: 10, 2022: 20, 2023: 15},
        'Outono': {2021: 5, 2022: 10, 2023: 8},
        'Inverno': {2021: 8, 2022: 12, 2023: 10},
        'Primavera': {2021: 15, 2022: 18, 2023: 20}
    }

    plot_grafico_comparativo(comparativo_por_estacao)

def calcular_consumo_por_estacao(dados):
    # Esta função calcula o consumo de energia por estação ao longo dos anos
    consumo_por_estacao = {}
    for row in dados.iter_rows(min_row=2, values_only=True):
        try:
            data_referencia = datetime.strptime(row[2], "%Y-%m-%d")
            ano = str(data_referencia.year)
            estacao = identificar_estacao(data_referencia.month)
            carga = row[5]
            if estacao not in consumo_por_estacao:
                consumo_por_estacao[estacao] = {}
            if ano not in consumo_por_estacao[estacao]:
                consumo_por_estacao[estacao][ano] = 0
            consumo_por_estacao[estacao][ano] += carga
        except ValueError:
            print(f"Erro ao processar a data {row[2]}")
    return consumo_por_estacao

def identificar_estacao(mes):
    # Esta função identifica a estação do ano com base no mês
    if mes in [12, 1, 2]:
        return 'Verão'
    elif mes in [3, 4, 5]:
        return 'Outono'
    elif mes in [6, 7, 8]:
        return 'Inverno'
    elif mes in [9, 10, 11]:
        return 'Primavera'

def encontrar_maior_consumo(consumo_por_estacao):
    # Esta função encontra a estação com o maior consumo de energia
    max_consumo = 0
    for estacao, consumo_por_ano in consumo_por_estacao.items():
        for ano, consumo in consumo_por_ano.items():
            if consumo > max_consumo:
                max_consumo = consumo
                estacao_maior_consumo = estacao
                ano_maior_consumo = ano
    return estacao_maior_consumo, ano_maior_consumo

def encontrar_menor_consumo(consumo_por_estacao):
    # Esta função encontra a estação com o menor consumo de energia
    min_consumo = float('inf')  # Inicializa o menor consumo com um valor infinito
    for estacao, consumo_por_ano in consumo_por_estacao.items():
        for ano, consumo in consumo_por_ano.items():
            if consumo < min_consumo:
                min_consumo = consumo
                estacao_menor_consumo = estacao
                ano_menor_consumo = ano
    return estacao_menor_consumo, ano_menor_consumo

def comparar_consumo_entre_anos(consumo_por_estacao):
    # Esta função compara o consumo de energia entre os anos por estação
    comparativo = {}
    for estacao, consumo_por_ano in consumo_por_estacao.items():
        anos = list(consumo_por_ano.keys())
        anos.sort()
        crescimento = consumo_por_ano[anos[-1]] - consumo_por_ano[anos[0]]
        comparativo[estacao] = crescimento
    return comparativo

def gerar_excel(resultados):
    # Esta função gera um arquivo Excel com os resultados
    wb = Workbook()
    ws = wb.active
    ws.append(['Estação', 'Crescimento'])
    for estacao, crescimento in resultados.items():
        ws.append([estacao, crescimento])
    wb.save('resultado_consumo.xlsx')

    # Carrega os dados do arquivo 'dados_api.xlsx' e calcula o consumo por estação
    wb_dados = openpyxl.load_workbook('dados_api.xlsx')
    ws_dados = wb_dados.active
    consumo_por_estacao = calcular_consumo_por_estacao(ws_dados)
    estacao_maior_consumo, ano_maior_consumo = encontrar_maior_consumo(consumo_por_estacao)
    print(f"A estação com maior consumo de carga foi {estacao_maior_consumo} no ano {ano_maior_consumo}")
    comparativo = comparar_consumo_entre_anos(consumo_por_estacao)
    
    # Chama a função gerar_excel novamente para criar o novo arquivo
    gerar_excel(comparativo)

    print("Resultados salvos em resultado_consumo.xlsx")

if __name__ == "__main__":
    main()